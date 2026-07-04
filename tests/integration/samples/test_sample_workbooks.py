from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


def _rows(path: Path) -> list[tuple[object, ...]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    return list(worksheet.iter_rows(values_only=True))


def test_required_sample_workbooks_exist_with_expected_contracts() -> None:
    project_root = Path(__file__).resolve().parents[3]
    samples = project_root / "samples"

    valid_rows = _rows(samples / "valid_import.xlsx")
    mixed_rows = _rows(samples / "mixed_import.xlsx")
    missing_header_rows = _rows(samples / "missing_headers_import.xlsx")
    duplicate_rows = _rows(samples / "duplicate_codes_import.xlsx")

    expected_headers = (
        "shipment_code",
        "customer_name",
        "origin_city",
        "destination_city",
        "weight_kg",
        "price",
        "status",
        "delivery_date",
    )

    assert valid_rows[0] == expected_headers
    assert mixed_rows[0] == expected_headers
    assert len(valid_rows) == 2
    assert len(mixed_rows) == 3

    assert "status" not in missing_header_rows[0]

    assert duplicate_rows[1][0] == duplicate_rows[2][0] == "SHP-DUP-1"
