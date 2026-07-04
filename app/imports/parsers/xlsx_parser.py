from __future__ import annotations

from collections.abc import Iterator
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook

from app.imports.parsers.errors import WorkbookStructureError
from app.imports.parsers.workbook_contract import REQUIRED_WORKBOOK_HEADERS, WORKBOOK_ROW_FIELDS

DEFAULT_ROW_CHUNK_SIZE = 500


@dataclass(frozen=True, slots=True)
class ParsedWorkbookRow:
    row_number: int
    values: dict[str, Any | None]


class XlsxParser:
    def __init__(self, workbook_path: Path, chunk_size: int = DEFAULT_ROW_CHUNK_SIZE) -> None:
        self._workbook_path = workbook_path
        self._chunk_size = chunk_size

    def iter_chunks(self) -> Iterator[list[ParsedWorkbookRow]]:
        workbook = None
        try:
            workbook = load_workbook(
                self._workbook_path,
                read_only=True,
                data_only=True,
            )
            worksheet = workbook.worksheets[0]
            headers = self._parse_headers(worksheet)

            rows: list[ParsedWorkbookRow] = []
            for row_number, row_values in self._iter_data_rows(worksheet, headers):
                rows.append(ParsedWorkbookRow(row_number=row_number, values=row_values))
                if len(rows) >= self._chunk_size:
                    yield rows
                    rows = []

            if rows:
                yield rows
        except (BadZipFile, OSError, ValueError, IndexError) as exc:
            raise WorkbookStructureError("Workbook structure could not be processed.") from exc
        finally:
            if workbook is not None:
                workbook.close()

    def _parse_headers(self, worksheet) -> dict[str, int]:
        header_values = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header_values is None:
            raise WorkbookStructureError("Workbook is missing a header row.")

        normalized_headers: dict[str, int] = {}
        for index, value in enumerate(header_values):
            if value is None:
                continue
            header = self._normalize_text(value)
            if header is None:
                continue
            if header in normalized_headers:
                raise WorkbookStructureError("Workbook contains duplicate headers.")
            normalized_headers[header] = index

        missing_headers = [
            header for header in REQUIRED_WORKBOOK_HEADERS if header not in normalized_headers
        ]
        if missing_headers:
            raise WorkbookStructureError(
                "Workbook is missing required headers: " + ", ".join(missing_headers)
            )

        return normalized_headers

    def _iter_data_rows(
        self,
        worksheet,
        headers: dict[str, int],
    ) -> Iterator[tuple[int, dict[str, Any | None]]]:
        for row_number, row_values in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            normalized_row = self._normalize_row(row_values, headers)
            if all(value is None for value in normalized_row.values()):
                continue

            yield row_number, normalized_row

    def _normalize_row(
        self,
        row_values: tuple[Any, ...],
        headers: dict[str, int],
    ) -> dict[str, Any | None]:
        normalized_row: dict[str, Any | None] = {}
        for field in WORKBOOK_ROW_FIELDS:
            index = headers.get(field)
            if index is None or index >= len(row_values):
                normalized_row[field] = None
                continue
            normalized_row[field] = self._normalize_text(row_values[index])
        for field in REQUIRED_WORKBOOK_HEADERS:
            normalized_row.setdefault(field, None)
        normalized_row.setdefault("delivery_date", None)
        return normalized_row

    @staticmethod
    def _normalize_text(value: object) -> Any | None:
        if isinstance(value, str):
            normalized = value.strip()
            return normalized or None
        return value
